"""Generate a formatted Excel workbook from site profile data."""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Profile data (mirrors src/data/site.ts)
SITE = {
    "name": "Jack Faris",
    "title": "Sales Leader & Robotics Specialist",
    "tagline": "Medtronic robotics specialist · AI builder",
    "location": "St. Louis, Missouri",
    "email": "jackafaris@gmail.com",
    "linkedin": "https://www.linkedin.com/in/jack-faris",
    "url": "https://jackfaris.com",
}

EXPERIENCE = [
    {
        "company": "Medtronic",
        "role": "Hugo RAS Platform Specialist",
        "period": "2024 – Present",
        "location": "St. Louis, MO",
        "highlights": [
            "Certified specialist on the Hugo Robotic-Assisted Surgery platform, supporting cases in the operating room.",
            "Demonstrate clinical and economic value to surgeons and healthcare systems evaluating robotic surgery programs.",
            "Serve as a trusted technical partner during case support, training, and program adoption.",
        ],
    },
    {
        "company": "Medical Device & Insurance Sales",
        "role": "Sales Professional",
        "period": "Pre-2024",
        "location": "Missouri",
        "highlights": [
            "Built a track record in B2B sales across medical devices and the insurance industry.",
            "Developed and executed strategies to exceed revenue goals and strengthen client relationships.",
            "Mentored peers and contributed to team development in fast-paced sales environments.",
        ],
    },
]

PROJECTS = [
    {
        "name": "OpenClaw",
        "period": "2024 – Present",
        "summary": "An AI agent platform designed to help small businesses and individuals automate real work with practical, accessible tools.",
        "highlights": [
            "Architected agent workflows, memory structures, and sub-agent patterns from the ground up.",
            "Evaluated multiple LLMs to match capability with real-world use cases.",
            "Focused on early adoption of AI to create durable value before the market saturates.",
        ],
    },
    {
        "name": "Medtronic Hugo RAS",
        "period": "2024 – Present",
        "summary": "Hands-on specialization in robotic-assisted surgery — bridging clinical teams, technology, and adoption.",
        "highlights": [
            "Platform-certified specialist supporting robotic surgery cases.",
            "Translates complex technology into outcomes surgeons and administrators care about.",
            "Represents the intersection of healthcare sales and deep product expertise.",
        ],
    },
]

TESTIMONIALS = [
    {
        "quote": "Jack brings rare clarity to complex technology. He connects clinical needs with the right solution and follows through in high-pressure environments.",
        "attribution": "Healthcare Partner",
        "context": "Robotic surgery program support",
    },
    {
        "quote": "He thinks in systems, asks sharp questions, and turns ideas into action. Jack is the kind of person you want driving a new initiative.",
        "attribution": "Former Colleague",
        "context": "Sales & team development",
    },
]

EDUCATION = {
    "school": "University of Missouri",
    "degree": "Bachelor of Health Science",
    "minor": "Sales Certificate Minor",
    "certifications": [
        "Medtronic Hugo RAS Platform Certified",
        "Health and Accident Insurance Producer License",
        "Excel Essential Training (Office 365)",
    ],
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def style_header_row(ws, row: int, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def auto_width(ws, min_width: int = 12, max_width: int = 60) -> None:
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        length = max(len(str(cell.value or "")) for cell in col_cells)
        ws.column_dimensions[letter].width = min(max(length + 2, min_width), max_width)


def build_workbook() -> Workbook:
    wb = Workbook()

    # --- Profile sheet ---
    ws = wb.active
    ws.title = "Profile"
    ws["A1"] = SITE["name"]
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:B1")

    profile_rows = [
        ("Title", SITE["title"]),
        ("Tagline", SITE["tagline"]),
        ("Location", SITE["location"]),
        ("Email", SITE["email"]),
        ("LinkedIn", SITE["linkedin"]),
        ("Website", SITE["url"]),
    ]
    for idx, (label, value) in enumerate(profile_rows, start=3):
        ws.cell(row=idx, column=1, value=label).font = Font(bold=True)
        ws.cell(row=idx, column=2, value=value)
        ws.cell(row=idx, column=1).border = THIN_BORDER
        ws.cell(row=idx, column=2).border = THIN_BORDER

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 55

    # --- Experience sheet ---
    ws_exp = wb.create_sheet("Experience")
    exp_headers = ["Company", "Role", "Period", "Location", "Highlights"]
    for col, header in enumerate(exp_headers, start=1):
        ws_exp.cell(row=1, column=col, value=header)
    style_header_row(ws_exp, 1, len(exp_headers))

    row = 2
    for item in EXPERIENCE:
        ws_exp.cell(row=row, column=1, value=item["company"])
        ws_exp.cell(row=row, column=2, value=item["role"])
        ws_exp.cell(row=row, column=3, value=item["period"])
        ws_exp.cell(row=row, column=4, value=item["location"])
        ws_exp.cell(row=row, column=5, value="\n".join(f"• {h}" for h in item["highlights"]))
        for col in range(1, 6):
            ws_exp.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="top")
            ws_exp.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    auto_width(ws_exp)
    ws_exp.column_dimensions["E"].width = 70

    # --- Projects sheet ---
    ws_proj = wb.create_sheet("Projects")
    proj_headers = ["Project", "Period", "Summary", "Highlights"]
    for col, header in enumerate(proj_headers, start=1):
        ws_proj.cell(row=1, column=col, value=header)
    style_header_row(ws_proj, 1, len(proj_headers))

    row = 2
    for item in PROJECTS:
        ws_proj.cell(row=row, column=1, value=item["name"])
        ws_proj.cell(row=row, column=2, value=item["period"])
        ws_proj.cell(row=row, column=3, value=item["summary"])
        ws_proj.cell(row=row, column=4, value="\n".join(f"• {h}" for h in item["highlights"]))
        for col in range(1, 5):
            ws_proj.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="top")
            ws_proj.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    auto_width(ws_proj)
    ws_proj.column_dimensions["C"].width = 55
    ws_proj.column_dimensions["D"].width = 65

    # --- Education sheet ---
    ws_edu = wb.create_sheet("Education")
    ws_edu["A1"] = "Education & Certifications"
    ws_edu["A1"].font = TITLE_FONT
    ws_edu.merge_cells("A1:B1")

    ws_edu["A3"] = "School"
    ws_edu["B3"] = EDUCATION["school"]
    ws_edu["A4"] = "Degree"
    ws_edu["B4"] = EDUCATION["degree"]
    ws_edu["A5"] = "Minor"
    ws_edu["B5"] = EDUCATION["minor"]
    for r in range(3, 6):
        ws_edu.cell(row=r, column=1).font = Font(bold=True)
        for c in range(1, 3):
            ws_edu.cell(row=r, column=c).border = THIN_BORDER

    ws_edu["A7"] = "Certifications"
    ws_edu["A7"].font = Font(bold=True)
    for idx, cert in enumerate(EDUCATION["certifications"], start=8):
        ws_edu.cell(row=idx, column=1, value=cert)
        ws_edu.cell(row=idx, column=1).border = THIN_BORDER

    ws_edu.column_dimensions["A"].width = 45
    ws_edu.column_dimensions["B"].width = 35

    # --- Testimonials sheet ---
    ws_test = wb.create_sheet("Testimonials")
    test_headers = ["Quote", "Attribution", "Context"]
    for col, header in enumerate(test_headers, start=1):
        ws_test.cell(row=1, column=col, value=header)
    style_header_row(ws_test, 1, len(test_headers))

    row = 2
    for item in TESTIMONIALS:
        ws_test.cell(row=row, column=1, value=item["quote"])
        ws_test.cell(row=row, column=2, value=item["attribution"])
        ws_test.cell(row=row, column=3, value=item["context"])
        for col in range(1, 4):
            ws_test.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="top")
            ws_test.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    auto_width(ws_test)
    ws_test.column_dimensions["A"].width = 70

    return wb


if __name__ == "__main__":
    output_path = "/workspace/Jack_Faris_Profile.xlsx"
    build_workbook().save(output_path)
    print(f"Created {output_path}")
